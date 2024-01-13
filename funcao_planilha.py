from openpyxl import Workbook
import xlwings as xw


def planilha(numero_processo, processo_anexado, lista_prestadora, lista_cnpj, lista_servico, lista_orgao, lista_vara, processo_principal,
             processo_associado, tipologia_processo, data_atuacao, data_liminar, doc_judicializacao, status_processo_administrativo, status_processo, data_consulta,
             regulamento_oficial, dispositivo_oficial, descricao_infracao, titulo, questionamento, tese, keywords,
             data_primeira_instancia, primeira_inclinacao):

    print("Criando novo arquivo .xlsx")
    wb = Workbook()
    ws = wb.active
    print("Arquivo .xlsx criado")
    # Criar as abas da planilha

    print("Criando abas do arquivo .xlsx")
    aba1 = wb.create_sheet('Status')
    aba2 = wb.create_sheet('Regulamento')
    aba3 = wb.create_sheet('Questionamentos')
    aba4 = wb.create_sheet('Chaves')
    aba5 = wb.create_sheet('Decisoes')

    worksheet = wb['Sheet']
    worksheet.title = 'Base'
    print("Abas do arquivo .xlsx criadas")

    print("Nomeando o cabeçalho de cada aba")
    worksheet.cell(row=1, column=1, value="Processo Administrativo")
    worksheet.cell(row=1, column=2, value="Processos Administrativo Anexados")
    worksheet.cell(row=1, column=3, value="Unidade Responsável")
    worksheet.cell(row=1, column=4, value="Prestadora")
    worksheet.cell(row=1, column=5, value="CNPJ")
    worksheet.cell(row=1, column=6, value="Serviço")
    worksheet.cell(row=1, column=7, value="Orgão Judicializado")
    worksheet.cell(row=1, column=8, value="Vara")
    worksheet.cell(row=1, column=9, value="Processo Judicial Principal")
    worksheet.cell(row=1, column=10, value="Processo Judicial Associado")
    worksheet.cell(row=1, column=11, value="Tipo Processo")
    worksheet.cell(row=1, column=12, value="Data de atuação")
    worksheet.cell(row=1, column=13, value="Data da Decisão Liminar")
    worksheet.cell(row=1, column=14, value="Documento de Judicialização")
    aba1.cell(row=1, column=1, value="Processo Administrativo")
    aba1.cell(row=1, column=2, value="Status do Processo no Judicial")
    aba1.cell(row=1, column=3, value="Data da Última Consulta")
    aba1.cell(row=1, column=4, value="Status do Processo Administrativo")
    aba2.cell(row=1, column=1, value="Processo Administrativo")
    aba2.cell(row=1, column=2, value="Regulamento")
    aba2.cell(row=1, column=3, value="Dispositivo")
    aba2.cell(row=1, column=4, value="Descrição da Infração")
    aba3.cell(row=1, column=1, value="Processo Administrativo")
    aba3.cell(row=1, column=2, value="Título dos Questionamentos")
    aba3.cell(row=1, column=3, value="Motivo padr x Questionamento padr.")
    aba3.cell(row=1, column=4, value="Questionamentos Padronizados")
    aba4.cell(row=1, column=1, value="Processo Administrativo")
    aba4.cell(row=1, column=2, value="Palavras-Chave")
    aba5.cell(row=1, column=1, value="processo")
    aba5.cell(row=1, column=2, value="motivo_da_judicializacao")
    aba5.cell(row=1, column=3, value="motivo_padronizado")
    aba5.cell(row=1, column=4, value="data_decisao_judicial_de_1_instancia")
    aba5.cell(row=1, column=5,
              value="inclinacao_da_decisao_judicial_de_1_instancia")
    aba5.cell(row=1, column=6, value="data_decisao_judicial_de_2_instancia")
    aba5.cell(row=1, column=7,
              value="inclinacao_da_decisao_judicial_de_2_instancia")
    print("Cabeçalho preenchido")

    print("Passando as informações para a aba Base")
    for i, string in enumerate(numero_processo, start=1):
        worksheet.cell(row=i+1, column=1).value = string
        aba1.cell(row=i+1, column=1).value = string
        aba2.cell(row=i+1, column=1).value = string
        aba4.cell(row=i+1, column=1).value = string
        worksheet.cell(row=i+1, column=3, value="CODI")

    for i, string in enumerate(processo_anexado, start=1):
        worksheet.cell(row=i+1, column=2).value = string

    for i, string in enumerate(lista_prestadora, start=1):
        worksheet.cell(row=i+1, column=4).value = string

    for i, string in enumerate(lista_cnpj, start=1):
        worksheet.cell(row=i+1, column=5).value = string

    for i, string in enumerate(lista_servico, start=1):
        worksheet.cell(row=i+1, column=6).value = string

    for i, string in enumerate(lista_orgao, start=1):
        worksheet.cell(row=i+1, column=7).value = string

    for i, string in enumerate(lista_vara, start=1):
        worksheet.cell(row=i+1, column=8).value = string

    for i, string in enumerate(processo_principal, start=1):
        worksheet.cell(row=i+1, column=9).value = string

    for i, string in enumerate(processo_associado, start=1):
        worksheet.cell(row=i+1, column=10).value = string

    for i, string in enumerate(tipologia_processo, start=1):
        worksheet.cell(row=i+1, column=11).value = string

    for i, string in enumerate(data_atuacao, start=1):
        worksheet.cell(row=i+1, column=12).value = string

    for i, string in enumerate(data_liminar, start=1):
        worksheet.cell(row=i+1, column=13).value = string

    for i, string in enumerate(doc_judicializacao, start=1):
        worksheet.cell(row=i+1, column=14).value = string
    print("Informações completas na aba Base")

    # Passando as infomações escolhidas para a aba "Status"
    print("Passando as informações para a aba Status")
    for i, string in enumerate(status_processo, start=1):
        aba1.cell(row=i+1, column=2).value = string

    for i, string in enumerate(data_consulta, start=1):
        aba1.cell(row=i+1, column=3).value = string

    for i, string in enumerate(status_processo_administrativo, start=1):
        aba1.cell(row=i+1, column=4).value = string
    print("Informações completas na aba Status")

    # Passando as informações escolhidas para a aba "Regulamento"
    print("Passando as informações para a aba Regulamento")
    row_index = 1
    item_index = 0

    # Loop para salvar cada título e número do processo em uma linha diferente
    # Trabalhando com sublist pois os regulamentos, dispositivos e descrições ficaram salvos em diferentes listas
    for sublist in regulamento_oficial:
        num = numero_processo[item_index]

        for string in sublist:
            aba2.cell(row=row_index+1, column=1).value = num
            aba2.cell(row=row_index+1, column=2).value = string
            row_index += 1

        item_index += 1

    row_index = 1
    for sublist in dispositivo_oficial:
        for string in sublist:
            aba2.cell(row=row_index+1, column=3).value = string
            row_index += 1

    row_index = 1
    for sublist in descricao_infracao:
        for string in sublist:
            aba2.cell(row=row_index+1, column=4).value = string
            row_index += 1
    print("Informações completas na aba Regulamento")

    # Passando as infomações escolhidas para a aba "Questionamentos"
    print("Passando as informações para a aba Questionamentos")
    row_index = 1
    item_index = 0

    # Loop para salvar cada título e número do processo em uma linha diferente
    # Trabalhando com sublist pois as teses, questionamentos e motivos ficaram salvos em diferentes listas
    for sublist in titulo:
        num = numero_processo[item_index]

        for string in sublist:
            aba3.cell(row=row_index+1, column=1).value = num
            aba3.cell(row=row_index+1, column=2).value = string
            row_index += 1

        item_index += 1

    row_index = 1
    for sublist in questionamento:
        for string in sublist:
            aba3.cell(row=row_index+1, column=3).value = string
            row_index += 1

    row_index = 1
    for sublist in tese:
        for string in sublist:
            aba3.cell(row=row_index+1, column=4).value = string
            row_index += 1
    print("Informações completas na aba Questionamentos")

    # Passando as infomações escolhidas para a aba "Chave"
    print("Passando as informações para a aba Chave")
    for i, string in enumerate(keywords, start=1):
        aba4.cell(row=i+1, column=2).value = string
    print("Informações completas na aba Chave")

    print("Passando as informações para a aba Decisoes")
    row_index = 1
    item_index = 0

    for sublist in titulo:
        num = numero_processo[item_index]
        data_inst = data_primeira_instancia[item_index]
        decisao_inclinacao_1 = primeira_inclinacao[item_index]

        for string in sublist:
            aba5.cell(row=row_index+1, column=1).value = num
            aba5.cell(row=row_index+1, column=2).value = string
            aba5.cell(row=row_index+1, column=4).value = data_inst
            aba5.cell(row=row_index+1, column=5).value = decisao_inclinacao_1
            row_index += 1

        item_index += 1

    row_index = 1
    for sublist in questionamento:
        for string in sublist:
            aba5.cell(row=row_index+1, column=3).value = string
            row_index += 1

    wb.save('Judicializacao_novas_fichas.xlsx')
